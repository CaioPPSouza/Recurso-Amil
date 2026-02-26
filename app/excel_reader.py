from __future__ import annotations

import csv
from pathlib import Path
import re
from typing import Dict, Iterable
import unicodedata

from openpyxl import load_workbook

from app.models import SpreadsheetRow

REQUIRED_COLUMNS = ("numero_guia", "senha", "valor_glosa", "justificativa")
OPTIONAL_COLUMNS = ("codigo_glosa",)
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "numero_guia": (
        "numero_da_guia_no_prestador",
        "numero_da_guia_atribuido_pela_operadora",
        "numero_guia",
        "numero_da_guia",
        "numero_guia_senha",
        "num_guia",
        "nr_guia",
        "n_guia",
        "guia",
    ),
    "senha": (
        "senha",
        "senha_da_guia",
        "senha_guia",
    ),
    "valor_glosa": (
        "valor_glosa",
        "valor_da_glosa",
        "valor_glosa_r",
        "valor_da_glosa_r",
        "valor_glosa_rs",
        "valor_glosado",
        "vl_glosa",
    ),
    "justificativa": (
        "justificativa_para_recurso",
        "justificativa",
        "justificativa_da_glosa",
        "justificativa_glosa",
        "motivo_glosa",
        "motivo",
    ),
    "codigo_glosa": (
        "codigo_da_glosa_da_guia",
        "codigo_glosa_da_guia",
        "codigo_da_glosa",
        "codigo_glosa",
        "cod_glosa",
    ),
}


def _normalize_header(name: str) -> str:
    if name is None:
        return ""
    text = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode(
        "ascii"
    )
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _parse_decimal(raw_value: object, line_number: int) -> float:
    if raw_value is None:
        raise ValueError(f"valor_glosa vazio na linha {line_number}")

    if isinstance(raw_value, (int, float)):
        return float(raw_value)

    text = str(raw_value).strip()
    if not text:
        raise ValueError(f"valor_glosa vazio na linha {line_number}")

    # Suporta "1.234,56" e "1234.56".
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(
            f"valor_glosa invalido na linha {line_number}: {raw_value!r}"
        ) from exc


def _build_header_mapping(headers: Iterable[str]) -> dict[str, str]:
    normalized = [header for header in headers if header]
    mapping: dict[str, str] = {}
    for canonical in (*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS):
        aliases = HEADER_ALIASES.get(canonical, (canonical,))
        found = next((alias for alias in aliases if alias in normalized), None)
        if found:
            mapping[canonical] = found

    missing = [column for column in REQUIRED_COLUMNS if column not in mapping]
    if missing:
        found_headers = ", ".join(normalized) if normalized else "<sem cabecalho>"
        raise ValueError(
            "Planilha com colunas obrigatorias ausentes. "
            f"Esperado: {', '.join(REQUIRED_COLUMNS)}. "
            f"Encontrado: {found_headers}"
        )
    return mapping


def _row_to_model(data: dict[str, object], line_number: int) -> SpreadsheetRow:
    numero_guia = str(data["numero_guia"]).strip()
    senha = str(data["senha"]).strip()
    justificativa = str(data["justificativa"]).strip()

    if not numero_guia or not senha:
        raise ValueError(f"numero_guia/senha vazios na linha {line_number}")
    if not justificativa:
        raise ValueError(f"justificativa vazia na linha {line_number}")

    codigo_glosa_raw = data.get("codigo_glosa")
    codigo_glosa = None
    if codigo_glosa_raw is not None:
        text = str(codigo_glosa_raw).strip()
        codigo_glosa = text or None

    return SpreadsheetRow(
        numero_guia=numero_guia,
        senha=senha,
        valor_glosa=_parse_decimal(data["valor_glosa"], line_number),
        justificativa=justificativa,
        codigo_glosa=codigo_glosa,
    )


def _read_csv_rows(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if not reader.fieldnames:
            raise ValueError("Arquivo CSV sem cabecalho.")
        headers = [_normalize_header(item) for item in reader.fieldnames]
        header_mapping = _build_header_mapping(headers)

        result = []
        for raw_row in reader:
            normalized_row = {
                _normalize_header(key): value for key, value in raw_row.items() if key
            }
            canonical_row = {
                canonical: normalized_row.get(source)
                for canonical, source in header_mapping.items()
            }
            result.append(canonical_row)
        return result


def _read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Planilha XLSX vazia.") from exc

    headers = [_normalize_header(item) for item in headers_row]
    header_mapping = _build_header_mapping(headers)

    parsed = []
    for values in rows:
        if not values or all(item is None for item in values):
            continue
        normalized_row = {headers[index]: value for index, value in enumerate(values)}
        canonical_row = {
            canonical: normalized_row.get(source)
            for canonical, source in header_mapping.items()
        }
        parsed.append(canonical_row)

    workbook.close()
    return parsed


def _read_rows(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix == ".xlsx":
        return _read_xlsx_rows(path)
    raise ValueError("Formato nao suportado. Use .csv ou .xlsx.")


def load_spreadsheet_index(path: Path) -> Dict[str, SpreadsheetRow]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    rows = _read_rows(path)
    index: Dict[str, SpreadsheetRow] = {}
    for line_number, row_data in enumerate(rows, start=2):
        model = _row_to_model(row_data, line_number)
        if model.key in index:
            raise ValueError(
                f"Planilha contem chave duplicada ({model.key}) na linha {line_number}."
            )
        index[model.key] = model

    return index
